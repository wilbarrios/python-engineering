# from openpyxl import Workbook # Create a workbook from scratch
from openpyxl import load_workbook

# wb = Workbook() # Create a workbook from scratch
wb = load_workbook(filename='example.xlsx')
ws = wb.active

ws['A1'] = 'Wilmer'
ws['D1'] = 'Barrios'

wb.save(filename='example.xlsx')
